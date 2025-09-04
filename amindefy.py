import os
from openpyxl import Workbook, load_workbook
from copy import copy

def amindefy_timesheets(timesheet_folder: str, output_file: str):
    """
    Combines Excel files into one workbook while preserving formatting.
    """
    # Create a new workbook
    output_wb = Workbook()
    output_wb.remove(output_wb.active)  # Remove default sheet
    
    for filename in os.listdir(timesheet_folder):
        if filename.endswith(".xlsx"):
            file_path = os.path.join(timesheet_folder, filename)
            
            # Load source workbook
            source_wb = load_workbook(file_path)
            source_ws = source_wb.active  # or specify sheet name
            
            # Create new sheet in output workbook
            sheet_name = os.path.splitext(filename)[0]
            output_ws = output_wb.create_sheet(title=sheet_name)
            
            # Copy all cells with formatting
            for row in source_ws.iter_rows():
                for cell in row:
                    new_cell = output_ws.cell(
                        row=cell.row, 
                        column=cell.column, 
                        value=cell.value
                    )
                    
                    # Copy formatting
                    if cell.has_style:
                        new_cell.font = copy(cell.font)
                        new_cell.border = copy(cell.border)
                        new_cell.fill = copy(cell.fill)
                        new_cell.number_format = cell.number_format
                        new_cell.protection = copy(cell.protection)
                        new_cell.alignment = copy(cell.alignment)

            # Copy column dimensions
            for col_letter, col_dimension in source_ws.column_dimensions.items():
                output_ws.column_dimensions[col_letter].width = col_dimension.width
            
            # Copy row dimensions
            for row_num, row_dimension in source_ws.row_dimensions.items():
                output_ws.row_dimensions[row_num].height = row_dimension.height
            
            source_wb.close()
    
    # Save the output workbook
    output_wb.save(output_file)
    output_wb.close()
    print(f"All timesheets have been combined into {output_file}.")